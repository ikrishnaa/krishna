from tkinter import *
from tkinter import ttk
import pandas as pd
import openpyxl
from openpyxl import Workbook
import pathlib
from ttkthemes import themed_tk as thtk
from tkinter import font
import numpy as np
from datetime import datetime

window = thtk.ThemedTk()
window.title("SHREE SAI COMPUTERS (CRM)")
window.geometry('1200x1200')
window.get_themes()
window.set_theme('ubuntu')
head = Label(window, text='SHREE SAI COMPUTERS', fg='#D27708', bg='#1A0FFA',
             font=font.Font(family='DotumChe', size=22, weight='bold', slant='italic', underline=0, overstrike=0))
head.pack(pady=0, fill=X)

defFont1 = font.Font(family='Courier', size=18, weight='bold', slant='roman', underline=0, overstrike=0)
font1 = font.Font(family='DotumChe', size=14, weight='bold', slant='italic', underline=0, overstrike=0)
today = datetime.now()
date = today.strftime("%d %b,%Y, %I:%M:%p")

# PENDING DB-------------------------------------------------
file = pathlib.Path('PENDING.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'DATE'
    sheet['B1'] = 'PRODUCT TIME'
    sheet['C1'] = 'PRODUCT INFO'
    sheet['D1'] = 'NAME'
    sheet['E1'] = 'CONTACT'
    sheet['F1'] = 'PROBLEM'
    sheet['G1'] = 'STATUS'
    sheet['H1'] = 'COST'
    sheet['F1'] = 'NOTE'

    file.save('PENDING.xlsx')

# Completed -------------------------------------------------
file1 = pathlib.Path('COMPLETED.xlsx')
if file1.exists():
    pass
else:
    file1 = Workbook()
    sheet['A1'] = 'DATE'
    sheet['B1'] = 'PRODUCT TIME'
    sheet['C1'] = 'PRODUCT INFO'
    sheet['D1'] = 'NAME'
    sheet['E1'] = 'CONTACT'
    sheet['F1'] = 'PROBLEM'
    sheet['G1'] = 'STATUS'
    sheet['H1'] = 'COST'
    sheet['F1'] = 'NOTE'

    file1.save('COMPLETED.xlsx')
# final DATABASE FILE ---------------------------------------
file2 = pathlib.Path('DATABASE.xlsx')
if file2.exists():
    pass
else:
    file2 = Workbook()
    sheet = file2.active
    sheet['A1'] = 'DATE'
    sheet['B1'] = 'PRODUCT TIME'
    sheet['C1'] = 'PRODUCT INFO'
    sheet['D1'] = 'NAME'
    sheet['E1'] = 'CONTACT'
    sheet['F1'] = 'PROBLEM'
    sheet['G1'] = 'STATUS'
    sheet['H1'] = 'COST'
    sheet['F1'] = 'NOTE'

    file2.save('DATABASE.xlsx')
# ORDER -------------------------------------------------------
file3 = pathlib.Path('ORDERS.xlsx')

if file3.exists():
    pass
else:
    file3 = Workbook()
    sheet = file3.active
    sheet['A1'] = 'DATE'
    sheet['B1'] = 'ITEM / PRODUCT'
    sheet['C1'] = 'NAME'
    sheet['D1'] = 'CONTACT'
    sheet['E1'] = 'TOTAL AMOUNT'
    sheet['F1'] = 'ADVANCE PAID'
    sheet['G1'] = 'REMAINING AMOUNT'
    sheet['H1'] = 'STATUS'

    file3.save('ORDERS.xlsx')


# -----------------------------------------------------------


def submit():
    ptype = ptypeval.get()
    pinfo = pinfoval.get()
    nval = nameval.get()
    cval = contactval.get()
    pval = problemval.get()
    ctval = costval.get()
    ntval = noteval.get()
    today = datetime.now()
    date = today.strftime("%d %b,%Y, %I:%M:%p")

    file = openpyxl.load_workbook('PENDING.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=date)
    sheet.cell(column=2, row=sheet.max_row, value=ptype)
    sheet.cell(column=3, row=sheet.max_row, value=pinfo)
    sheet.cell(column=4, row=sheet.max_row, value=nval)
    sheet.cell(column=5, row=sheet.max_row, value=cval)
    sheet.cell(column=6, row=sheet.max_row, value=pval)
    sheet.cell(column=7, row=sheet.max_row, value='PENDING')
    sheet.cell(column=8, row=sheet.max_row, value=ctval)
    sheet.cell(column=9, row=sheet.max_row, value=ntval)

    file.save('PENDING.xlsx')

    # e1.delete(0, END)
    e2.delete(0, END)
    e3.delete(0, END)
    e4.delete(0, END)
    e5.delete(0, END)
    e6.delete(0, END)
    e7.delete(0, END)
    refresh()


def cancel():
    e2.delete(0, END)
    e3.delete(0, END)
    e4.delete(0, END)
    e5.delete(0, END)
    e6.delete(0, END)
    e7.delete(0, END)


notebook = ttk.Notebook(window)
notebook.pack(pady=15)

frame1 = Frame(notebook, width=900, height=700, bg='white')
frame2 = Frame(notebook, width=900, height=700, bg='white')
frame3 = Frame(notebook, width=900, height=700, bg='white')
frame4 = Frame(notebook, width=900, height=700, bg='white')
frame5 = Frame(notebook, width=900, height=700, bg='white')
frame6 = Frame(notebook, width=900, height=700, bg='white')
frame7 = Frame(notebook, width=900, height=700, bg='white')

frame1.pack()
frame2.pack()
frame3.pack()
frame4.pack()
frame5.pack()
frame6.pack()
frame7.pack()

notebook.add(frame1, text='NEW')
notebook.add(frame2, text='PENDING')
notebook.add(frame3, text='COMPLETED')
notebook.add(frame4, text='ORDERS')
notebook.add(frame5, text='MARKET DETAILS')
notebook.add(frame6, text='MARKET LIST')
notebook.add(frame7, text='CUSTOMERS')

# FRAME 1 NEW ------------------------

product_type = Label(frame1, text='PRODUCT TYPE', font=font1, bg='white')
product_type.grid(row=1, column=1, padx=10, pady=10)
product_info = Label(frame1, text='PRODUCT INFO', font=font1, bg='white')
product_info.grid(row=2, column=1, padx=10, pady=10)
name = Label(frame1, text='NAME ', font=font1, bg='white')
name.grid(row=3, column=1, padx=10, pady=10)
contact = Label(frame1, text="CONTACT", font=font1, bg='white')
contact.grid(row=4, column=1, padx=10, pady=10)
problem = Label(frame1, text='PROBLEM', font=font1, bg='white')
problem.grid(row=5, column=1, padx=10, pady=10)
cost = Label(frame1, text='ESTIMATED COST', font=font1, bg='white')
cost.grid(row=6, column=1, padx=10, pady=10)
note = Label(frame1, text='NOTE', font=font1, bg='white')
note.grid(row=7, column=1, padx=10, pady=10)

Label(frame1, text='* mandatory', fg='red', bg='white').grid(row=4, column=3)

# ptypeval = StringVar()
# e1 = Entry(frame1, width=30, textvariable=ptypeval)
options = ['Laptop', 'Laptop with Adapter', 'PC', 'Printer', 'Mobile']
ptypeval = StringVar()
ptypeval.set('---SELECT OPTION---')
drop = OptionMenu(frame1, ptypeval, *options)
drop.config(width=30)

pinfoval = StringVar()
e2 = Entry(frame1, width=30, textvariable=pinfoval)
nameval = StringVar()
e3 = Entry(frame1, width=30, textvariable=nameval)
contactval = StringVar(value=np.nan)
e4 = Entry(frame1, width=30, textvariable=contactval)
problemval = StringVar()
e5 = Entry(frame1, width=30, textvariable=problemval)
costval = StringVar()
e6 = Entry(frame1, width=30, textvariable=costval)
noteval = StringVar()
e7 = Entry(frame1, width=30, textvariable=noteval)

drop.grid(row=1, column=2, padx=10, pady=10)
e2.grid(row=2, column=2, padx=10, pady=10)
e3.grid(row=3, column=2, padx=10, pady=10)
e4.grid(row=4, column=2, padx=10, pady=10)
e5.grid(row=5, column=2, padx=10, pady=10)
e6.grid(row=6, column=2, padx=10, pady=10)
e7.grid(row=7, column=2, padx=10, pady=10)

b1 = Button(frame1, text='Submit ↩ ', font=defFont1, bg='white', borderwidth=5, command=submit)
b1.grid(row=10, column=2, padx=0, pady=0)

b2 = Button(frame1, text='Clear ⌫', font=defFont1, bg='white', borderwidth=5, command=cancel)
b2.grid(row=10, column=3, padx=10, pady=10)

close = Button(frame1, text='Close ⎋', font=defFont1, bg='white', borderwidth=5, command=window.quit)
close.grid(row=10, column=4, padx=10, pady=10)

# FRAME 2 PENDING----------------------

framep = Frame(frame2, width=900, height=500)
frameedit = Frame(frame2, width=700, height=300)

scroll = Scrollbar(framep)

table = ttk.Treeview(framep, yscrollcommand=scroll.set)
scroll.pack(side=RIGHT, fill=Y)
scroll.config(command=table.yview)

style = ttk.Style(table)
style.theme_use('clam')
style.configure('', font=('Helvetica', 11))
style.configure('Treeview.Heading', foreground='red', font=('Helvetica', 13, 'bold'))
style.configure('Trerview.Column', foreground='red', font=('Helvetica', 13, 'bold'))
framep.pack(pady=20)

lb1 = Label(frameedit, text='P Type')
lb1.grid(row=1, column=1)
lb2 = Label(frameedit, text='P Info')
lb2.grid(row=1, column=2)
lb3 = Label(frameedit, text='Name')
lb3.grid(row=1, column=3)
lb4 = Label(frameedit, text='contact')
lb4.grid(row=1, column=4)
lb5 = Label(frameedit, text='Problem')
lb5.grid(row=1, column=5)
lb6 = Label(frameedit, text='Status')
lb6.grid(row=1, column=6)
lb7 = Label(frameedit, text='Cost')
lb7.grid(row=1, column=7)
lb8 = Label(frameedit, text='Note')
lb8.grid(row=1, column=8)

q = Entry(frameedit, width=10)
q.grid(row=2, column=1)
q1 = Entry(frameedit, width=10)
q1.grid(row=2, column=2)
q2 = Entry(frameedit, width=10)
q2.grid(row=2, column=3)
q3 = Entry(frameedit, width=10)
q3.grid(row=2, column=4)
q4 = Entry(frameedit, width=10)
q4.grid(row=2, column=5)
q5 = Entry(frameedit, width=10)
q5.grid(row=2, column=6)
q6 = Entry(frameedit, width=10)
q6.grid(row=2, column=7)
q7 = Entry(frameedit, width=10)
q7.grid(row=2, column=8)

frameedit.pack()

file5 = r"{}".format('PENDING.xlsx')
df = pd.read_excel(file5)

table['column'] = list(['DATE', 'PRODUCT TYPE', 'PRODUCT INFO', 'NAME', 'CONTACT', 'PROBLEM', 'STATUS', 'COST', 'NOTE'])
table['show'] = 'headings'

for col in table['column']:
    table.column(col, width=120, anchor=CENTER)
    table.heading(col, text=col)
df_rows = df.to_numpy().tolist()
for row in df_rows:
    table.insert("", 'end', value=row[0:9])

table.pack(padx=10)


def editp():
    q.delete(0, END)
    q1.delete(0, END)
    q2.delete(0, END)
    q3.delete(0, END)
    q4.delete(0, END)
    q5.delete(0, END)
    q6.delete(0, END)
    q7.delete(0, END)

    selected = table.focus()
    global key
    key = table.item(selected, 'values')

    q.insert(0, key[1])
    q1.insert(0, key[2])
    q2.insert(0, key[3])
    q3.insert(0, key[4])
    q4.insert(0, key[5])
    q5.insert(0, key[6])
    q6.insert(0, key[7])
    q7.insert(0, key[8])


def updatep():
    selected = table.focus()
    table.item(selected, text='', values=(q.get(), q1.get(), q2.get(), q3.get(), q4.get(), q5.get(), q6.get()))
    file8 = openpyxl.load_workbook('PENDING.xlsx')
    sheet2 = file8.active
    count = 1

    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[4] == key[4]:
            print('cell', cell[4])
            print('key', key[4])
            sheet2.delete_rows(count)
    file8.save('PENDING.xlsx')

    file = openpyxl.load_workbook('PENDING.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=date)
    sheet.cell(column=2, row=sheet.max_row, value=q.get())
    sheet.cell(column=3, row=sheet.max_row, value=q1.get())
    sheet.cell(column=4, row=sheet.max_row, value=q2.get())
    sheet.cell(column=5, row=sheet.max_row, value=q3.get())
    sheet.cell(column=6, row=sheet.max_row, value=q4.get())
    sheet.cell(column=7, row=sheet.max_row, value=q5.get())
    sheet.cell(column=8, row=sheet.max_row, value=q6.get())
    sheet.cell(column=9, row=sheet.max_row, value=q7.get())

    file.save('PENDING.xlsx')

    q.delete(0, END)
    q1.delete(0, END)
    q2.delete(0, END)
    q3.delete(0, END)
    q4.delete(0, END)
    q5.delete(0, END)
    q6.delete(0, END)
    q7.delete(0, END)
    refresh()


def refresh():
    clear()
    file5 = r"{}".format('PENDING.xlsx')
    df = pd.read_excel(file5)

    table['column'] = list(
        ['DATE', 'PRODUCT TYPE', 'PRODUCT INFO', 'NAME', 'CONTACT', 'PROBLEM', 'STATUS', 'COST', 'NOTE'])
    table['show'] = 'headings'

    for col2 in table['column']:
        table.column(col2, width=120, anchor=CENTER)
        table.heading(col2, text=col2)
    df_rows = df.to_numpy().tolist()
    for row in df_rows:
        table.insert("", 'end', value=row[0:9])


def clear():
    table.delete(*table.get_children())


def complete():
    s = table.focus()
    v = table.item(s, 'values')
    x = table.selection()[0]
    table.delete(x)

    file4 = openpyxl.load_workbook('COMPLETED.xlsx')
    sheet1 = file4.active

    sheet1.cell(column=1, row=sheet1.max_row + 1, value=v[0])
    sheet1.cell(column=2, row=sheet1.max_row, value=v[1])
    sheet1.cell(column=3, row=sheet1.max_row, value=v[2])
    sheet1.cell(column=4, row=sheet1.max_row, value=v[3])
    sheet1.cell(column=5, row=sheet1.max_row, value=v[4])
    sheet1.cell(column=6, row=sheet1.max_row, value=v[5])
    sheet1.cell(column=7, row=sheet1.max_row, value='Not Delivered')
    sheet1.cell(column=8, row=sheet1.max_row, value=v[7])
    sheet1.cell(column=9, row=sheet1.max_row, value=v[8])

    file4.save('COMPLETED.xlsx')

    file8 = openpyxl.load_workbook('PENDING.xlsx')
    sheet2 = file8.active
    count = 1
    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[3] == v[3]:
            sheet2.delete_rows(count)

    file8.save('PENDING.xlsx')
    print(cell[3], v[3])
    completed_ref()


def completed_ref():
    table1.delete(*table1.get_children())
    file6 = r"{}".format('COMPLETED.xlsx')
    df1 = pd.read_excel(file6)

    table1['column'] = list(
        ['DATE', 'PRODUCT TYPE', 'PRODUCT INFO', 'NAME', 'CONTACT', 'PROBLEM', 'STATUS', 'COST', 'NOTE'])
    table1['show'] = 'headings'

    for col3 in table1['column']:
        table1.column(col3, width=120, anchor=CENTER)
        table1.heading(col3, text=col3)
    df_rows1 = df1.to_numpy().tolist()
    for row1 in df_rows1:
        table1.insert("", 'end', value=row1[0:9])


def erase_row():
    s = table.focus()
    v1 = table.item(s, 'values')
    x = table.selection()[0]
    table.delete(x)

    file8 = openpyxl.load_workbook('PENDING.xlsx')
    sheet2 = file8.active
    count = 1
    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[3] == v1[3]:
            sheet2.delete_rows(count)

    file8.save('PENDING.xlsx')


ref_btn = Button(frame2, text='Refresh', font=defFont1, bg='white', borderwidth=5, command=refresh)
# ref_btn.pack(pady=10)

done_btn = Button(frame2, text='Checked ✓', font=defFont1, bg='white', borderwidth=5, command=complete)
done_btn.pack(pady=10)

erase = Button(frame2, text='Delete ✗ ', font=defFont1, bg='white', borderwidth=5, command=erase_row)
erase.pack(pady=10)

edit_btn = Button(frame2, text='Edit ✂ ', font=defFont1, bg='white', borderwidth=5, command=editp)
edit_btn.pack(pady=10)

update_btn = Button(frame2, text='Update ⇡', font=defFont1, bg='white', borderwidth=5, command=updatep)
update_btn.pack(pady=10)
# FRAME 3 COMPLETE-------------------------

framec = Frame(frame3, width=900, height=500)
framecom = Frame(frame3, width=700, height=300)

scroll1 = Scrollbar(framec)
table1 = ttk.Treeview(framec, yscrollcommand=scroll1.set)

scroll1.pack(side=RIGHT, fill=Y)
scroll1.config(command=table1.yview)

style = ttk.Style(table)
style.configure('Treeview', rowheight=30)
framec.pack(pady=20)

llb1 = Label(framecom, text='P Type')
llb1.grid(row=1, column=1)
llb2 = Label(framecom, text='P Info')
llb2.grid(row=1, column=2)
llb3 = Label(framecom, text='Name')
llb3.grid(row=1, column=3)
llb4 = Label(framecom, text='contact')
llb4.grid(row=1, column=4)
llb5 = Label(framecom, text='Problem')
llb5.grid(row=1, column=5)
llb6 = Label(framecom, text='Status')
llb6.grid(row=1, column=6)
llb7 = Label(framecom, text='Cost')
llb7.grid(row=1, column=7)
llb8 = Label(framecom, text='Note')
llb8.grid(row=1, column=8)

qn = Entry(framecom, width=10)
qn.grid(row=2, column=1)
qn1 = Entry(framecom, width=10)
qn1.grid(row=2, column=2)
qn2 = Entry(framecom, width=10)
qn2.grid(row=2, column=3)
qn3 = Entry(framecom, width=10)
qn3.grid(row=2, column=4)
qn4 = Entry(framecom, width=10)
qn4.grid(row=2, column=5)
qn5 = Entry(framecom, width=10)
qn5.grid(row=2, column=6)
qn6 = Entry(framecom, width=10)
qn6.grid(row=2, column=7)
qn7 = Entry(framecom, width=10)
qn7.grid(row=2, column=8)

framecom.pack()

file7 = r"{}".format('COMPLETED.xlsx')
df1 = pd.read_excel(file7)

table1['column'] = list(
    ['DATE', 'PRODUCT TYPE', 'PRODUCT INFO', 'NAME', 'CONTACT', 'PROBLEM', 'STATUS', 'COST', 'NOTE'])
table1['show'] = 'headings'

for col4 in table['column']:
    table1.column(col4, width=120, anchor=CENTER)
    table1.heading(col4, text=col4)

df_rows1 = df1.to_numpy().tolist()
for row1 in df_rows1:
    table1.insert("", 'end', value=row1[0:9])

table1.pack(padx=10)


def editc():
    qn.delete(0, END)
    qn1.delete(0, END)
    qn2.delete(0, END)
    qn3.delete(0, END)
    qn4.delete(0, END)
    qn5.delete(0, END)
    qn6.delete(0, END)
    qn7.delete(0, END)

    selected = table1.focus()
    global key1
    key1 = table1.item(selected, 'values')

    qn.insert(0, key1[1])
    qn1.insert(0, key1[2])
    qn2.insert(0, key1[3])
    qn3.insert(0, key1[4])
    qn4.insert(0, key1[5])
    qn5.insert(0, key1[6])
    qn6.insert(0, key1[7])
    qn7.insert(0, key1[8])


def updatec():
    selected = table1.focus()
    table1.item(selected, text='', values=(qn.get(), qn1.get(), qn2.get(), qn3.get(), qn4.get(), qn5.get(), qn6.get()))
    file8 = openpyxl.load_workbook('COMPLETED.xlsx')
    sheet2 = file8.active
    count = 1
    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[3] == key1[3]:
            sheet2.delete_rows(count)
    file8.save('COMPLETED.xlsx')

    file = openpyxl.load_workbook('COMPLETED.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=date)
    sheet.cell(column=2, row=sheet.max_row, value=qn.get())
    sheet.cell(column=3, row=sheet.max_row, value=qn1.get())
    sheet.cell(column=4, row=sheet.max_row, value=qn2.get())
    sheet.cell(column=5, row=sheet.max_row, value=qn3.get())
    sheet.cell(column=6, row=sheet.max_row, value=qn4.get())
    sheet.cell(column=7, row=sheet.max_row, value=qn5.get())
    sheet.cell(column=8, row=sheet.max_row, value=qn6.get())
    sheet.cell(column=9, row=sheet.max_row, value=qn7.get())

    file.save('COMPLETED.xlsx')

    qn.delete(0, END)
    qn1.delete(0, END)
    qn2.delete(0, END)
    qn3.delete(0, END)
    qn4.delete(0, END)
    qn5.delete(0, END)
    qn6.delete(0, END)
    qn7.delete(0, END)
    completed_ref()


def backtop():
    s1 = table1.focus()
    v1 = table1.item(s1, 'values')
    q = table1.selection()[0]
    table1.delete(q)

    file3 = openpyxl.load_workbook('PENDING.xlsx')
    sheet = file3.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=v1[0])
    sheet.cell(column=2, row=sheet.max_row, value=v1[1])
    sheet.cell(column=3, row=sheet.max_row, value=v1[2])
    sheet.cell(column=4, row=sheet.max_row, value=v1[3])
    sheet.cell(column=5, row=sheet.max_row, value=v1[4])
    sheet.cell(column=6, row=sheet.max_row, value=v1[5])
    sheet.cell(column=7, row=sheet.max_row, value='PENDING')
    sheet.cell(column=8, row=sheet.max_row, value=v1[7])
    sheet.cell(column=9, row=sheet.max_row, value=v1[8])

    file3.save('PENDING.xlsx')

    file8 = openpyxl.load_workbook('COMPLETED.xlsx')
    sheet2 = file8.active
    count = 1
    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[3] == v1[3]:
            sheet2.delete_rows(count)

    file8.save('COMPLETED.xlsx')
    print(cell[3], v1[3])
    refresh()


def deliver():
    s1 = table1.focus()
    v1 = table1.item(s1, 'values')
    q = table1.selection()[0]
    table1.delete(q)

    file3 = openpyxl.load_workbook('DATABASE.xlsx')
    sheet = file3.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=v1[0])
    sheet.cell(column=2, row=sheet.max_row, value=v1[1])
    sheet.cell(column=3, row=sheet.max_row, value=v1[2])
    sheet.cell(column=4, row=sheet.max_row, value=v1[3])
    sheet.cell(column=5, row=sheet.max_row, value=v1[4])
    sheet.cell(column=6, row=sheet.max_row, value=v1[5])
    sheet.cell(column=7, row=sheet.max_row, value='REPAIRED')
    sheet.cell(column=8, row=sheet.max_row, value=v1[7])
    sheet.cell(column=9, row=sheet.max_row, value=v1[8])

    file3.save('DATABASE.xlsx')

    file8 = openpyxl.load_workbook('COMPLETED.xlsx')
    sheet2 = file8.active
    count = 1
    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[3] == v1[3]:
            sheet2.delete_rows(count)

    file8.save('COMPLETED.xlsx')
    print(cell[3], v1[3])
    database_ref()


def returned():
    s1 = table1.focus()
    v1 = table1.item(s1, 'values')
    q = table1.selection()[0]
    table1.delete(q)
    file3 = openpyxl.load_workbook('DATABASE.xlsx')
    sheet = file3.active

    sheet.cell(column=1, row=sheet.max_row + 1, value=v1[0])
    sheet.cell(column=2, row=sheet.max_row, value=v1[1])
    sheet.cell(column=3, row=sheet.max_row, value=v1[2])
    sheet.cell(column=4, row=sheet.max_row, value=v1[3])
    sheet.cell(column=5, row=sheet.max_row, value=v1[4])
    sheet.cell(column=6, row=sheet.max_row, value=v1[5])
    sheet.cell(column=7, row=sheet.max_row, value="Not Repaired")
    sheet.cell(column=8, row=sheet.max_row, value=v1[7])
    sheet.cell(column=9, row=sheet.max_row, value=v1[8])

    file3.save('DATABASE.xlsx')

    file9 = openpyxl.load_workbook('COMPLETED.xlsx')
    sheet2 = file9.active
    count = 1
    for cell in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        if cell[3] == v1[3]:
            sheet2.delete_rows(count)

    file9.save('COMPLETED.xlsx')
    database_ref()


# com_ref = Button(frame3, text='Refresh', font=defFont1, bg='white', borderwidth=5, command=completed_ref)
# com_ref.pack(pady=10)

backtopend = Button(frame3, text='← Pending', font=defFont1, bg='white', borderwidth=5, command=backtop)
backtopend.pack(pady=10)

del_btn = Button(frame3, text='Deliver ✓', font=defFont1, bg='white', borderwidth=5, command=deliver)
del_btn.pack(pady=10)

ret_btn = Button(frame3, text='Returned ⇢', font=defFont1, bg='white', borderwidth=5, command=returned)
ret_btn.pack(pady=10)

edit2_btn = Button(frame3, text='Edit ✂', font=defFont1, bg='white', borderwidth=5, command=editc)
edit2_btn.pack(pady=10)

update2_btn = Button(frame3, text='Update ⇡', font=defFont1, bg='white', borderwidth=5, command=updatec)
update2_btn.pack(pady=10)


# FRAME 4 ORDER----------------------------

def submit_order():
    it = itemval.get()
    nm = mnameval.get()
    con = mcontactval.get()
    tot = totalval.get()
    adv = advamountval.get()
    today = datetime.now()

    file = openpyxl.load_workbook('ORDERS.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=date)
    sheet.cell(column=2, row=sheet.max_row, value=it)
    sheet.cell(column=3, row=sheet.max_row, value=nm)
    sheet.cell(column=4, row=sheet.max_row, value=con)
    sheet.cell(column=5, row=sheet.max_row, value=tot)
    sheet.cell(column=6, row=sheet.max_row, value=adv)
    sheet.cell(column=7, row=sheet.max_row, value=int(tot) - int(adv))
    sheet.cell(column=8, row=sheet.max_row, value='Pending')

    file.save('ORDERS.xlsx')
    print(it)

    o1.delete(0, END)
    o2.delete(0, END)
    o3.delete(0, END)
    o4.delete(0, END)
    o5.delete(0, END)
    order_ref()
    market_ref()


def clear_order():
    o1.delete(0, END)
    o2.delete(0, END)
    o3.delete(0, END)
    o4.delete(0, END)
    o5.delete(0, END)


item1 = Label(frame4, text='ITEM / PRODUCT', font=font1, bg='white')
item1.grid(row=1, column=1, padx=10, pady=10)
mname = Label(frame4, text='NAME', font=font1, bg='white')
mname.grid(row=2, column=1, padx=10, pady=10)
mcontact = Label(frame4, text='CONTACT', font=font1, bg='white')
mcontact.grid(row=3, column=1, padx=10, pady=10)
total = Label(frame4, text='TOTAL AMOUNT', font=font1, bg='white')
total.grid(row=4, column=1, padx=10, pady=10)
advamount = Label(frame4, text='ADVANCE AMOUNT', font=font1, bg='white')
advamount.grid(row=5, column=1, padx=10, pady=10)

itemval = StringVar()
o1 = Entry(frame4, width=30, textvariable=itemval)
mnameval = StringVar()
o2 = Entry(frame4, width=30, textvariable=mnameval)
mcontactval = StringVar()
o3 = Entry(frame4, width=30, textvariable=mcontactval)
totalval = StringVar()
o4 = Entry(frame4, width=30, textvariable=totalval)
advamountval = StringVar()
o5 = Entry(frame4, width=30, textvariable=advamountval)

order_button = Button(frame4, text='Place Orde ⏎ ', font=defFont1, bg='white', borderwidth=5, command=submit_order)
cancel_order = Button(frame4, text='Clear Fields ⌫', font=defFont1, bg='white', borderwidth=5, command=clear_order)
order_button.grid(row=10, column=2, padx=10, pady=10)
cancel_order.grid(row=10, column=3, padx=10, pady=10)

o1.grid(row=1, column=2, padx=10, pady=10)
o2.grid(row=2, column=2, padx=10, pady=10)
o3.grid(row=3, column=2, padx=10, pady=10)
o4.grid(row=4, column=2, padx=10, pady=10)
o5.grid(row=5, column=2, padx=10, pady=10)


# FRAME 6 MARKET LIST ----------------------

def market_ref():
    table2.delete(*table2.get_children())

    file6 = r"{}".format('ORDERS.xlsx')
    df2 = pd.read_excel(file6)

    table2['column'] = list(['MARKET LIST'])
    table2['show'] = 'headings'

    for col5 in table2['column']:
        table2.column(col5, width=300, anchor=CENTER)
        table2.heading(col5, text=col5)
    df_rows2 = df2.to_numpy().tolist()
    for row2 in df_rows2:
        if row2[7] == 'Pending':
            table2.insert("", 'end', value=row2[1])


framem = Frame(frame6, width=500, height=500)

scroll2 = Scrollbar(framem)

table2 = ttk.Treeview(framem, yscrollcommand=scroll2.set)
scroll2.pack(side=RIGHT, fill=Y)
scroll2.config(command=table2.yview)

style1 = ttk.Style(table2)
style1.configure('Treeview', rowheight=30)
style1.configure('Treeview.Heading', foreground='red', font=('Helvetica', 13, 'bold'))
style1.configure('Trerview.Column', foreground='red', font=('Helvetica', 13, 'bold'))

framem.pack(pady=20)

file6 = r"{}".format('ORDERS.xlsx')
df2 = pd.read_excel(file6)

table2['column'] = list(['MARKET LIST'])
table2['show'] = 'headings'

for col6 in table2['column']:
    table2.column(col6, width=250, anchor=CENTER)
    table2.heading(col6, text=col6)
df_rows2 = df2.to_numpy().tolist()
for row2 in df_rows2:
    if row2[7] == 'Pending':
        table2.insert("", 'end', value=row2[1])

table2.pack(padx=10)

lref = Button(frame6, text='Refresh', command=market_ref, font=defFont1, bg='white', borderwidth=5)


# lref.pack(pady=10)
# FRAME 7 DATABASE----------------------------------------

def database_ref():
    table3.delete(*table3.get_children())

    file14 = r"{}".format('DATABASE.xlsx')
    df4 = pd.read_excel(file14)

    table3['column'] = list(
        ['DATE', 'PRODUCT TYPE', 'PRODUCT INFO', 'NAME', 'CONTACT', 'PROBLEM', 'STATUS', 'COST', 'NOTE'])
    table3['show'] = 'headings'

    for col7 in table3['column']:
        table3.column(col7, width=120, anchor=CENTER)
        table3.heading(col7, text=col7)

    df_rows4 = df4.to_numpy().tolist()
    for row4 in df_rows4:
        table3.insert("", 'end', value=row4[0:9])


framed = Frame(frame7, width=500, height=500)

scroll4 = Scrollbar(framed)
table3 = ttk.Treeview(framed, yscrollcommand=scroll4.set)

scroll4.pack(side=RIGHT, fill=Y)
scroll4.config(command=table3.yview)

style2 = ttk.Style(table3)
style2.configure('Treeview', rowheight=30)
style2.theme_use('clam')
style2.configure('', font=('Helvetica', 11))
style2.configure('Treeview.Heading', foreground='red', font=('Helvetica', 13, 'bold'))
framed.pack(pady=20)

file13 = r"{}".format('DATABASE.xlsx')
df4 = pd.read_excel(file13)

table3['column'] = list(
    ['DATE', 'PRODUCT TYPE', 'PRODUCT INFO', 'NAME', 'CONTACT', 'PROBLEM', 'STATUS', 'COST', 'NOTE'])
table3['show'] = 'headings'

for col8 in table3['column']:
    table3.column(col8, width=120, anchor=CENTER)
    table3.heading(col8, text=col8)

df_rows4 = df4.to_numpy().tolist()
for row4 in df_rows4:
    table3.insert("", 'end', value=row4[0:9])

table3.pack(padx=10)

dataref = Button(frame7, text='Refresh', font=defFont1, bg='white', borderwidth=5, command=database_ref)


# dataref.pack(pady=20)


# FRAME 5 MARKET DETAILS----------------------

def complete_order():
    file16 = r"{}".format('ORDERS.xlsx')
    df6 = pd.read_excel(file6)
    s1 = table4.focus()
    v1 = table4.item(s1, 'values')
    file12 = openpyxl.load_workbook('ORDERS.xlsx')
    sheet3 = file12.active
    count = 1
    for cell in sheet3.iter_rows(min_row=2, max_row=sheet3.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        # print('cell 2',cell[2],'cell 3',cell[3])
        if cell[3] == v1[3]:
            sheet3.cell(column=8, row=count, value='Complete')

    file12.save('ORDERS.xlsx')
    q = table4.selection()[0]
    table4.delete(q)
    order_ref()
    market_ref()


def order_ref():
    table4.delete(*table4.get_children())

    file16 = r"{}".format('ORDERS.xlsx')
    df6 = pd.read_excel(file6)

    table4['column'] = list(df6.columns)[0:-1]
    table4['show'] = 'headings'

    for col15 in table4['column']:
        table4.column(col15, width=170, anchor=CENTER)
        table4.heading(col15, text=col15)
    df_rows12 = df6.to_numpy().tolist()
    for row12 in df_rows12:
        if row12[7] == 'Pending':
            table4.insert("", 'end', value=row12)


def delete_order():
    s1 = table4.focus()
    v1 = table4.item(s1, 'values')
    file12 = openpyxl.load_workbook('ORDERS.xlsx')
    sheet3 = file12.active
    count = 1
    for cell in sheet3.iter_rows(min_row=2, max_row=sheet3.max_row, min_col=1, max_col=9, values_only=True):
        count += 1
        print('cell', cell[2], 'v', v1[2])
        if cell[2] == v1[2]:
            sheet3.delete_rows(count)

    file12.save('ORDERS.xlsx')
    q = table4.selection()[0]
    table4.delete(q)
    order_ref()
    market_ref()


framel = Frame(frame5, width=500, height=500)

scroll5 = Scrollbar(framel)
table4 = ttk.Treeview(framel, yscrollcommand=scroll5.set)

scroll5.pack(side=RIGHT, fill=Y)
scroll5.config(command=table4.yview)

style3 = ttk.Style(table4)
style3.configure('Treeview', rowheight=30)
style3.theme_use('clam')
style3.configure('', font=('Helvetica', 11))
style3.configure('Treeview.Heading', foreground='red', font=('Helvetica', 13, 'bold'))
framel.pack(pady=20)

file15 = r"{}".format('ORDERS.xlsx')
df5 = pd.read_excel(file15)

table4['column'] = list(df5.columns)[0:-1]
table4['show'] = 'headings'

for col9 in table4['column']:
    table4.column(col9, width=170, anchor=CENTER)
    table4.heading(col9, text=col9)

df_rows5 = df5.to_numpy().tolist()
for row5 in df_rows5:
    if row5[7] == 'Pending':
        table4.insert("", 'end', value=row5)

table4.pack(padx=10)

ref = Button(frame5, text='Refresh', font=defFont1, bg='white', borderwidth=5, command=order_ref)
# ref.pack(pady=20)

comp_order = Button(frame5, text='Complete Order ✔', font=defFont1, bg='white', borderwidth=5, command=complete_order)
comp_order.pack(pady=20)

del_order = Button(frame5, text='Cancel Order ✗', font=defFont1, bg='white', borderwidth=5, command=delete_order)
del_order.pack(pady=20)

tag = Label(window, text='Designed by Krishna Rajpurohit', anchor='se',
            font=font.Font(family='Gisha', size=18, weight='bold', slant='italic', underline=0, overstrike=0))
tag.pack()
window.mainloop()
